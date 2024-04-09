from openpyxl import load_workbook

# 打开工作簿
wb = load_workbook(r"E:\code\table_handle\data\drilling\1.xlsx")
sheet = wb["table"]
sheet1 = wb["fact verification"]

# 转换为 JSON 格式

propositions = []
for row in list(sheet1.iter_rows())[1:]:
    print(row[0].value)
    print(row[1].value)
