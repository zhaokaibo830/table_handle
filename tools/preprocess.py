import json
import math
from typing import Dict, List, Tuple, Any
# import cv2
# from PIL import Image
# import pytesseract
from openpyxl import load_workbook
import os
import zipfile
import xml.etree.ElementTree as ET
from xml.dom import minidom
from openpyxl.workbook import Workbook


def detect_text(image_path):
    # 读取图像
    image = cv2.imread(image_path)
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    # 二值化处理
    _, thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)

    # 查找轮廓
    contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    # 筛选出文本区域
    text_regions = []
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        # 过滤掉过小的区域
        if w > 1 and h > 1:
            text_regions.append((x, y, x + w, y + h))

    # 按照文本框的顶部位置进行排序
    text_regions.sort(key=lambda x: x[1])

    return text_regions


def draw_text_boxes(image_path, text_regions):
    # 读取图像
    image = cv2.imread(image_path)

    # 绘制文本框
    for (x1, y1, x2, y2) in text_regions:
        cv2.rectangle(image, (x1, y1), (x2, y2), (0, 255, 0), 2)

    # 保存带有文本框的图像
    output_image_path = image_path.replace('.png', '_text_boxes.png')
    cv2.imwrite(output_image_path, image)

    return output_image_path


def read_json_from_file(json_file_path: str) -> dict:
    with open(json_file_path, 'r', encoding='utf-8') as file:
        json_data = json.load(file)
    return json_data


def image_to_excel(image_path: str, excel_path: str):
    """
    Converts an image containing a table to Excel format.

    Args:
        image_path (str): The path to the image file.
        excel_path (str): The path to save the generated Excel file.
    """
    # 读取文本框并填充单元格
    text_regions = detect_text(image_path)

    # 计算 max_cols
    max_cols = int(math.sqrt(len(text_regions))) + 1

    # 创建一个新的 Excel 工作簿
    wb = Workbook()
    ws = wb.active

    # 遍历每个文本框并填充单元格，从第二个文本框开始
    for i, text_region in enumerate(text_regions[1:], start=1):  # 从第二个文本框开始遍历
        x1, y1, x2, y2 = text_region

        # 识别文本框中的内容
        with Image.open(image_path) as img:
            cropped_img = img.crop((x1, y1, x2, y2))
            cell_text = pytesseract.image_to_string(cropped_img, lang='chi_sim')

        # 移除文本中的换行符
        cell_text = cell_text.replace('\n', ' ')

        # 将文本填充到对应的单元格中
        ws.cell(row=(i - 1) // max_cols + 1, column=(i - 1) % max_cols + 1, value=cell_text.strip())

    # 保存 Excel 文件
    wb.save(excel_path)

    # 返回带有文本框的图像路径
    return draw_text_boxes(image_path, text_regions)


def extract_table_data_from_xml(xml_path: str) -> List[Dict[str, Any]]:
    """
    :param xml_path:
    :return cells:
    从解压的xml文件里提取信息，并转化成对应的json格式
    """
    # 解析 XML 文件
    tree = ET.parse(xml_path)
    root = tree.getroot()

    # 存储单元格数据的列表
    cells = []

    prev_col_end_idx = 0

    # 查找包含表格数据的节点
    for table in root.iter('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}tbl'):
        # 遍历表格中的行
        for row_idx, row in enumerate(table.findall('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}tr'),
                                      start=1):
            # 重置列索引为1
            col_idx = 1

            # 遍历行中的单元格
            for cell_idx, cell in enumerate(
                    row.findall('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}tc'), start=1):
                # 计算单元格的行跨度和列跨度
                row_span = 1
                col_span = 1

                # 获取单元格文本内容
                cell_text = ''
                for p in cell.findall('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}p'):
                    for r in p.findall('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}r'):
                        for t in r.findall('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t'):
                            if t.text is not None:
                                cell_text += t.text

                # 获取单元格的列合并信息（如果存在）
                tc_pr = cell.find('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}tcPr')
                if tc_pr is not None:
                    grid_span = tc_pr.find('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}gridSpan')
                    if grid_span is not None:
                        val = grid_span.attrib.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val')
                        if val is not None:
                            col_span = int(val)

                # 获取单元格的行合并信息（如果存在）
                v_merge = cell.find('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}tcPr/'
                                    '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}vMerge')

                if v_merge is not None:
                    # 初始化合并计数器
                    merge_counter = 2
                    # 向下搜索，直到找到不再存在行合并的单元格为止
                    for j in range(row_idx + 1, len(table.findall(
                            '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}tr')) + 1):
                        next_row_cell = \
                            table.findall('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}tr')[
                                j - 1].find(
                                '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}tc')
                        next_v_merge = next_row_cell.find(
                            '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}tcPr/'
                            '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}vMerge')
                        if next_v_merge is not None:
                            next_v_merge_text = next_v_merge.attrib.get(
                                '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val', '')
                            if next_v_merge_text == 'continue':
                                # 计数器自增
                                merge_counter += 1
                            else:
                                break
                        else:
                            break
                    # 计算合并的行数
                    row_span = merge_counter
                else:
                    row_span = 1

                # 判断单元格颜色，用于区分表头和表体
                cell_color = cell.find('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}tcPr/'
                                       '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}shd')
                if cell_color is not None:
                    cell_fill = cell_color.attrib.get(
                        '{http://schemas.openxmlformats.org/wordprocessingml/2006/main}fill')
                    if cell_fill == "C8C8C8":
                        node_type = "key"  # 表头
                    else:
                        node_type = "value"  # 表体
                else:
                    node_type = "value"  # 默认为表体

                # 构造单元格数据字典
                if cell_text.strip() != '':
                    cells.append({
                        "colspan": [col_idx, col_idx + col_span - 1],
                        "rowspan": [row_idx, row_idx + row_span - 1],
                        "text": cell_text.strip(),
                        "node_type": node_type
                    })

                # 更新前一个单元格的结束列索引
                prev_col_end_idx = col_idx + col_span - 1

                # 更新列索引为当前单元格的结束列索引加一
                col_idx = prev_col_end_idx + 1

    return cells


def format_xml(xml_path: str):
    """
    :param xml_path:
    :return:
    由于解压的xml默认在一行显示，所以这个函数用于格式化xml文件，便于阅读
    """
    # 读取 XML 文件内容
    with open(xml_path, "r", encoding="utf-8") as f:
        xml_content = f.read()

    # 解析 XML 内容
    dom = minidom.parseString(xml_content)

    # 格式化 XML 内容
    formatted_xml = dom.toprettyxml(indent="  ")

    # 写回 XML 文件
    with open(xml_path, "w", encoding="utf-8") as f:
        f.write(formatted_xml)


def excel_to_json(excel_path: str) -> Tuple:
    """
    输入一个excel文件，输出对应的json统一格式描述，excel是xlsx格式
    :return:
    """
    # 读取 Excel 文件
    wb = load_workbook(excel_path)
    sheet = wb["table"]
    sheet1 = wb["fact verification"]

    # 转换为 JSON 格式
    cells = []

    # 遍历每个单元格
    for row in sheet.iter_rows():
        for cell in row:
            # 获取单元格的值
            cell_value = cell.value

            # 检查单元格是否为空
            if cell_value is None:
                continue

            # 初始化单元格的跨度
            col_start = cell.column
            col_span = [col_start, col_start]
            row_span = [cell.row, cell.row]

            # 判断单元格是否为合并单元格
            for merged_cell_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_cell_range:
                    # 获取合并单元格的起始行号和列号
                    merged_start_cell = merged_cell_range.min_row, merged_cell_range.min_col
                    merged_end_cell = merged_cell_range.max_row, merged_cell_range.max_col
                    row_span = [merged_start_cell[0], merged_end_cell[0]]
                    col_span = [merged_start_cell[1], merged_end_cell[1]]
                    break

            # 检查单元格的颜色是否为特定的 RGB（200，200，200）
            cell_color = cell.fill.start_color.rgb
            if cell_color == "FFC8C8C8":
                node_type = "key"
            else:
                node_type = "value"

            # 添加单元格到列表中
            cells.append({
                "colspan": col_span,
                "rowspan": row_span,
                "text": str(cell_value).strip().replace("\n", " "),
                "node_type": node_type
            })

    propositions = []
    for row in list(sheet1.iter_rows())[1:]:
        propositions.append({"proposition": row[0].value, "value": row[1].value})
        # 输出 JSON 文件
        # with open('output.json', 'w', encoding='utf-8') as f:
        #     f.write("{\n")
        #     f.write('  "cells": [\n')
        #     for i, cell in enumerate(cells):
        #         f.write('    {\n')
        #         f.write(f'      "colspan": [{cell["colspan"][0]}, {cell["colspan"][1]}],\n')
        #         f.write(f'      "rowspan": [{cell["rowspan"][0]}, {cell["rowspan"][1]}],\n')
        #         f.write(f'      "text": "{cell["text"].strip().replace("\n", " ")}",\n')
        #         f.write(f'      "node_type": "{cell["node_type"]}"\n')
        #         f.write('    }')
        #         if i != len(cells) - 1:
        #             f.write(',')
        #         f.write('\n')
        #     f.write('  ]\n')
        #     f.write("}\n")
        #
        # # 读取生成的 JSON 文件并返回
        # json_data = read_json_from_file('output.json')
        # # print(json_data)
    return {"cells": cells}, propositions


def image_to_json(image_path: str) -> Dict:
    """
    输入一张表格图片，输出对应的json统一格式描述

    Args:
        image_path (str): 图像文件路径

    Returns:
        Dict: 包含表格信息的 JSON 对象
    """
    # 定义临时 Excel 文件路径
    temp_excel_path = "temp.xlsx"

    # 将图像转换为 Excel
    image_to_excel(image_path, temp_excel_path)

    # 将 Excel 转换为 JSON
    json_data,_ = excel_to_json(temp_excel_path)

    return json_data


def word_to_json(word_path: str) -> Dict:
    """
    输入包含一个表格的word文件，输出对应的json统一格式描述。word是docx格式
    :return:
    """
    # 解压 DOCX 文件
    with zipfile.ZipFile(docx_file, 'r') as zip_ref:
        zip_ref.extractall("temp")

    # 找到解压后的 XML 文件路径
    xml_files = [file for file in os.listdir("temp/word") if file.endswith('.xml')]
    xml_path = os.path.join("temp/word", xml_files[0])

    # 格式化 XML 文件
    format_xml(xml_path)

    # 提取表格数据
    table_data = extract_table_data_from_xml(xml_path)

    # 写入 JSON 文件
    with open('output.json', 'w', encoding='utf-8') as f:
        f.write("{\n")
        f.write('  "cells": [\n')
        for i, cell in enumerate(table_data):
            f.write('    {\n')
            f.write(f'      "colspan": [{cell["colspan"][0]}, {cell["colspan"][1]}],\n')
            f.write(f'      "rowspan": [{cell["rowspan"][0]}, {cell["rowspan"][1]}],\n')
            f.write(f'      "text": "{cell["text"]}",\n')
            f.write(f'      "node_type": "{cell["node_type"]}"\n')
            f.write('    }')
            if i != len(table_data) - 1:
                f.write(',')
            f.write('\n')
        f.write('  ]\n')
        f.write("}\n")

    # 读取生成的 JSON 文件并返回
    json_data = read_json_from_file('output.json')
    # print(json_data)
    return json_data


def any_format_to_json(file_path: str) -> Dict:
    """
    输入一个带有表格信息的xlsx/png/jpeg/jpg/docx文件，输出对应的json统一格式描述
    :param file_path: 文件路径
    :return: 表格信息的JSON描述
    """
    # 获取文件扩展名
    file_extension = file_path.split('.')[-1].lower()

    # 根据文件类型调用相应的函数
    if file_extension in ['png', 'jpeg', 'jpg']:
        return image_to_json(file_path)
    elif file_extension == 'xlsx':
        return excel_to_json(file_path)
    elif file_extension == 'docx':
        return word_to_json(file_path)
    else:
        raise ValueError("Unsupported file format. Only xlsx, png, jpeg, jpg, and docx are supported.")


if __name__ == '__main__':
    image_path = r"D:\project\torchlearing\ceshi\table.png"
    excel_path = r"E:\code\table_handle\data\player\2.xlsx"
    docx_file = "D:\\project\\torchlearing\\ceshi\\test.docx"
    output_json_path = "output.json"
    excel_to_json(excel_path)
    # word_to_json(docx_file)
    # image_to_json(image_path)
    # any_format_to_json(docx_file)
